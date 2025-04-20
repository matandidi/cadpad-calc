import tkinter as tk  # ייבוא מודול tkinter לממשק גרפי
from tkinter import messagebox, scrolledtext  # ייבוא תיבות הודעה ושדות גלילה
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg  # שילוב גרפי matplotlib ב-Tkinter
import matplotlib.pyplot as plt  # ייבוא pyplot לציור גרפים
from PIL import Image, ImageTk  # ייבוא PIL לטיפול בתמונות ב-Tkinter
from PIL import ImageDraw  # ייבוא ImageDraw לציור בתמונות
import datetime  # ייבוא מודול datetime לעבודה עם תאריכים
import os  # ייבוא מודול os לעבודה עם קבצים ותיקיות
import io  # ייבוא io לטיפול בבuffers בזיכרון

from openpyxl import Workbook  # ייבוא Workbook ליצירת קובצי Excel
from openpyxl.drawing.image import Image as OpxImage  # ייבוא OpxImage להטמעת תמונות ב-Excel

def safe_int(val):  # פונקציה להמרת מחרוזת למספר שלם בטוח
    val = val.strip()  # הסרת רווחים מהקצוות
    if val == "":  # בדיקת מחרוזת ריקה
        return 0  # החזרת 0 אם המחרוזת ריקה
    return int(val)  # המרת המחרוזת למספר שלם

def safe_float(val):  # פונקציה להמרת מחרוזת למספר עשרוני בטוח
    val = val.strip()  # הסרת רווחים מהקצוות
    if val == "":  # בדיקת מחרוזת ריקה
        return 0.0  # החזרת 0.0 אם המחרוזת ריקה
    return float(val)  # המרת המחרוזת למספר עשרוני

def select_all(event):  # פונקציה לבחור את כל הטקסט בשדה כניסה בעת מוקד
    event.widget.select_range(0, 'end')  # סימון כל התוכן
    event.widget.icursor('end')  # העתקת הסמן לסוף

def export_to_excel(masach_val, total_procurement, steps_log, leftover_by_year, years, fig_all):  # פונקציה לייצוא תוצאות לאקסל
    """  # התחלת תיאור הפונקציה
    מייצא:  # תיאור ראשי: ייצוא נתונים לאקסל
    1) מסח"א  # הוספת שדה MASACH
    2) כמות רכש סופית  # הוספת שדה הסתכמות הרכש
    3) לוג החישוב (steps_log)  # הוספת לוג מפורט
    4) טבלת leftover_by_year  # הוספת טבלת שאריות לפי שנה
    5) הגרף (fig_all) מוטמע בגליון האקסל (ללא קובץ PNG חיצוני)  # הטמעת גרף
    שם הקובץ בפורמט: ddMMyyyy_<masach_val>.xlsx  # פורמט שם הקובץ
    """  # סיום תיאור הפונקציה

    today_str = datetime.date.today().strftime("%d%m%Y")  # יצירת מחרוזת תאריך של היום
    short_id = f"{masach_val}_{today_str}"  # יצירת מזהה קצר משילוב מאסח ותאריך
    out_dir = 'exports'  # תיקיית היעד לשמירת הקבצים
    if not os.path.exists(out_dir):  # בדיקה אם התיקייה קיימת
        os.makedirs(out_dir)  # יצירת התיקייה אם אינה קיימת
    filename = os.path.join(out_dir, f"{short_id}.xlsx")  # הרכבת נתיב הקובץ החדש

    wb = Workbook()  # יצירת חוברת אקסל חדשה
    ws = wb.active  # בחירת הגיליון הפעיל הראשון
    ws.title = "CalculationResult"  # הגדרת שם הגיליון לתוצאה

    ws["A1"] = "MASACH"  # הגדרת כותרת MASACH בשורה 1, עמודה A
    ws["B1"] = masach_val  # הכנסת ערך MASACH לתא B1

    ws["A2"] = "Final Procurement Needed"  # כותרת עמודת כמות רכש סופית בשורה 2, עמודה A
    ws["B2"] = total_procurement  # הכנסת ערך סכום הרכש הסופי לתא B2

    # leftover table  # התחלת יצירת טבלת שאריות לפי שנה
    ws["E1"] = "Year"  # כותרת עמודת שנה עבור טבלת השאריות
    ws["F1"] = "Leftover"  # כותרת עמודת שאריות עבור טבלת השאריות
    row_idx = 2  # התחלת מילוי נתוני טבלת השאריות משורה 2
    for y in years:  # לולאה על כל השנים להגדרת טבלת השאריות
        ws.cell(row=row_idx, column=5, value=y)  # הכנסת השנה בעמודה E
        ws.cell(row=row_idx, column=6, value=leftover_by_year[y])  # הכנסת ערך השאריות בעמודה F
        row_idx += 1  # מעבר לשורה הבאה בטבלת השאריות

    # steps log  # התחלת כתיבת לוג שלב אחר שלב
    ws["A4"] = "Steps Log:"  # כותרת לוג השלבים בתא A4
    row_log = 5  # התחלת כתיבת לוג משורה 5
    for line in steps_log:  # לולאה על כל השורות בלוג
        ws.cell(row=row_log, column=1, value=line)  # הכנסת שורת לוג לעמודה A
        row_log += 1  # מעבר לשורה הבאה בלוג השלבים

    
    # --- Embed matplotlib figure in memory (no external PNG)  # התחלת יצירת תמונה בזיכרון ללא קובץ חיצוני ---
    img_buffer = io.BytesIO()  # יצירת buffer בזיכרון לאחסון תמונה
    fig_all.savefig(img_buffer, format="png", dpi=300, bbox_inches="tight")  # שמירת הגרף ל-buffer כ-PNG עם DPI גבוה ומעט שוליים
    img_buffer.seek(0)  # החזרת מצביע ה-buffer לתחילתו לקריאה
    
    pil_img = Image.open(img_buffer)  # פתיחת התמונה מתוך ה-buffer כ-PIL Image
    img = OpxImage(pil_img)  # המרת PIL Image לאובייקט תמונה עבור Excel
    img.width = 600  # קביעת רוחב התמונה בפיקסלים
    img.height = 400  # קביעת גובה התמונה בפיקסלים
    # הנחת התמונה בשורות מתחת לכל התוכן כדי שלא תכסה תוכן קיים בגליון
    last_row = ws.max_row  # קבלת מספר השורה האחרונה עם תוכן בגליון
    start_row = last_row + 2  # חישוב שורת העוגן עם שתי שורות ריקות מעל
    anchor_cell = f"A{start_row}"  # כתובת התא לעיגון התמונה
    ws.add_image(img, anchor_cell)  # הדבקת התמונה בגליון בתא שצויין
    
    try:  # ניסיון לשמירת חוברת ה-Excel לקובץ
        wb.save(filename)  # שמירת חוברת ה-Excel למיקום הקובץ
        messagebox.showinfo("Export to Excel", f"נשמר בהצלחה: {filename}")  # הודעת הצלחה למשתמש
    except Exception as e:  # טיפול בשגיאה בשמירה
        messagebox.showerror("Export Error", f"שגיאה בשמירה: {e}")  # הצגת הודעת שגיאה

def calculate_gap():
    try:
        masach_val = entry_masach.get()
        item_desc_val = entry_item_desc.get()
        nsn_val = entry_nsn.get()
        part_number_val = entry_part_number.get()
        cage_val = entry_cage.get()

        alt_parts = [e.get() for e in alt_entries]
        atud_factor = safe_float(entry_atud_factor.get())
        life_test_qty = safe_int(entry_life_test_qty.get())

        start_year = safe_int(entry_start_year.get())
        if start_year <= 0:
            raise ValueError("שנת ההתחלה חייבת להיות מספר חיובי")

        years = list(range(start_year, start_year + 9))

        shelf_life_storage_months = safe_int(entry_storage.get())
        shelf_life_use_months = safe_int(entry_use.get())

        extra_shelf_year = safe_int(entry_extra_shelf_year.get())
        extra_shelf_qty  = safe_int(entry_extra_shelf_qty.get())
        shelf_groups = []

        if extra_shelf_qty > 0 and extra_shelf_year > 0:
            year_expire_month = extra_shelf_year * 12
            year_produced_month = year_expire_month - shelf_life_storage_months
            group_extra = {
                'year_expire': extra_shelf_year,
                'year_expire_month': year_expire_month,
                'year_produced_month': year_produced_month,
                'quantity': extra_shelf_qty
            }
            shelf_groups.append(group_extra)

        installed_need_list = [safe_int(e.get()) for e in installed_entries]
        shelf_data_list     = [safe_int(e.get()) for e in shelf_entries]

        total_installed_sum = sum(installed_need_list)

        # Build shelf groups from the 9-year table
        for i, y in enumerate(years):
            qty = shelf_data_list[i]
            if qty > 0:
                year_expire_month = y * 12
                year_produced_month = year_expire_month - shelf_life_storage_months
                group = {
                    'year_expire': y,
                    'year_expire_month': year_expire_month,
                    'year_produced_month': year_produced_month,
                    'quantity': qty
                }
                shelf_groups.append(group)

        installed_need_by_year = {}
        for i, y in enumerate(years):
            installed_need_by_year[y] = installed_need_list[i]

        leftover_by_year = {y: 0 for y in years}
        steps_log = []

        def can_use_this_group(install_year, group):
            return group['year_expire'] >= (install_year + 2)

        for i_year in years:
            needed = installed_need_by_year.get(i_year, 0)
            leftover = needed

            if leftover <= 0:
                steps_log.append(f"Year {i_year}: no demand (0).")
                leftover_by_year[i_year] = 0
                continue

            steps_log.append(f"\n*** Year {i_year}, Demand = {needed} ***")
            steps_log.append(
                f"(Info) Storage life = {shelf_life_storage_months} months, "
                f"In-service life = {shelf_life_use_months} months, leftoverUse >= 24 (2yrs)."
            )

            shelf_groups.sort(key=lambda g: g['year_expire'])
            install_year_month = i_year * 12

            for group in shelf_groups:
                if leftover <= 0:
                    break
                if group['quantity'] <= 0:
                    continue
                if not can_use_this_group(i_year, group):
                    continue

                age_in_storage_months = install_year_month - group['year_produced_month']
                if age_in_storage_months < 0:
                    continue
                if age_in_storage_months >= shelf_life_storage_months:
                    continue

                leftover_in_storage_months = shelf_life_storage_months - age_in_storage_months
                leftover_use_months = min(leftover_in_storage_months, shelf_life_use_months)

                if leftover_use_months < 24:
                    continue

                used = min(leftover, group['quantity'])
                group['quantity'] -= used
                leftover -= used

                future_expire_month = install_year_month + leftover_use_months
                future_expire_year = future_expire_month // 12
                installed_need_by_year[future_expire_year] = installed_need_by_year.get(future_expire_year, 0) + used

                steps_log.append(
                    f"Year {i_year}: took {used} from (expire={group['year_expire']}), remain={group['quantity']}.\n"
                    f"ageInStorage={age_in_storage_months}, leftoverUse={leftover_use_months}, "
                    f"final expire={future_expire_year}, need={leftover}."
                )

            leftover_by_year[i_year] = leftover
            if leftover > 0:
                steps_log.append(f"Year {i_year}: leftover gap = {leftover}.")
            else:
                steps_log.append(f"Year {i_year}: fully covered, gap=0.")

        total_gap = sum(leftover_by_year[y] for y in years)
        steps_log.append("\n===========================")
        steps_log.append(f"Total gap = {total_gap}")
        steps_log.append(f"Sum of installed (input) = {total_installed_sum}")
        steps_log.append(f"Atud Factor = {atud_factor}")

        # Deal table
        total_future_orders = 0
        transaction_log = []
        for i in range(3):
            desc = deals_desc_entries[i].get()
            qty_str = deals_qty_entries[i].get()
            arrival_info = deals_arrival_entries[i].get()

            qty_order = safe_int(qty_str)
            if qty_order > 0:
                total_future_orders += qty_order

            if desc.strip() or qty_order > 0 or arrival_info.strip():
                transaction_log.append(f"Deal #{i+1}: desc={desc}, qty={qty_order}, arrival={arrival_info}")

        steps_log.append("\n--- Future deals table ---")
        steps_log.append(f"Total future orders quantity = {total_future_orders}")
        for line in transaction_log:
            steps_log.append(line)

        first_calc = total_installed_sum * atud_factor
        second_calc = total_gap - total_future_orders + life_test_qty
        total_procurement = first_calc + second_calc

        steps_log.append("\n--- Procurement Calculation ---")
        steps_log.append(f"First calc = (Sum installed) * (Atud factor) = {total_installed_sum} * {atud_factor} = {first_calc:.2f}")
        steps_log.append(f"Second calc = (gap) - (future orders) + life_test_qty = {total_gap} - {total_future_orders} + {life_test_qty} = {second_calc:.2f}")
        steps_log.append(f"Final procurement = {total_procurement:.2f}")

        # Result window
        result_window = tk.Toplevel(root)
        result_window.title("Gap Calculation Results")
        result_window.protocol("WM_DELETE_WINDOW", on_closing)

        top_frame = tk.Frame(result_window)
        top_frame.pack(side="top", fill="x", padx=10, pady=5)

        tk.Label(top_frame, text=f"MASACH: {masach_val}").pack(anchor="w")
        tk.Label(top_frame, text=f"Item Desc: {item_desc_val}").pack(anchor="w")
        tk.Label(top_frame, text=f"NSN: {nsn_val}").pack(anchor="w")
        tk.Label(top_frame, text=f"Part Number: {part_number_val}").pack(anchor="w")
        tk.Label(top_frame, text=f"CAGE: {cage_val}").pack(anchor="w")

        for i, alt in enumerate(alt_parts):
            alt_stripped = alt.strip()
            if alt_stripped:
                tk.Label(top_frame, text=f"Alternative #{i+1}: {alt_stripped}").pack(anchor="w")

        tk.Label(top_frame, text=f"Atud Factor: {atud_factor}").pack(anchor="w")
        tk.Label(top_frame, text=f"Life Test Qty: {life_test_qty}").pack(anchor="w")

        tk.Label(top_frame, text=f"Final Procurement Needed: {total_procurement:.2f}",
                 font=("Arial", 13, "bold"), fg="blue").pack(anchor="w", pady=5)

        def on_export_click():
            nonlocal fig_all, leftover_by_year, years
            export_to_excel(
                masach_val,
                total_procurement,
                steps_log,
                leftover_by_year,
                years,
                fig_all
            )

        header_frame = tk.Frame(result_window)
        header_frame.pack(side="left", padx=10, pady=10, anchor="n")

        tk.Label(header_frame, text="Year", width=10, borderwidth=1, relief="solid").grid(row=0, column=0)
        tk.Label(header_frame, text="Gap (Leftover)", width=15, borderwidth=1, relief="solid").grid(row=0, column=1)

        row_index = 1
        for y in years:
            tk.Label(header_frame, text=str(y), width=10, borderwidth=1, relief="solid").grid(row=row_index, column=0)
            tk.Label(header_frame, text=str(leftover_by_year[y]), width=15, borderwidth=1, relief="solid").grid(row=row_index, column=1)
            row_index += 1

        tk.Label(header_frame, text=f"Total gap: {total_gap}",
                 font=("Arial", 12, "bold"), pady=10).grid(row=row_index, column=0, columnspan=2)
        row_index += 1
        tk.Label(header_frame, text=f"Sum installed: {total_installed_sum}",
                 font=("Arial", 12, "bold"), pady=10).grid(row=row_index, column=0, columnspan=2)

        steps_frame = tk.Frame(result_window)
        steps_frame.pack(side="right", padx=10, pady=10, expand=True, fill="both")

        tk.Label(steps_frame, text="Steps Log", font=("Arial", 12, "bold")).pack()
        text_area = scrolledtext.ScrolledText(steps_frame, width=80, height=25, wrap="word")
        text_area.pack(expand=True, fill="both")
        for line in steps_log:
            text_area.insert("end", line + "\n")
        text_area.config(state="disabled")

        # build the figure
        fig_all, ax_all = plt.subplots(figsize=(10, 6))
        all_years_str = [str(y) for y in years]
        base_demand_all = [installed_need_by_year.get(y, 0) - leftover_by_year.get(y, 0) for y in years]
        injected_all = [0 for _ in years]
        covered_all = base_demand_all[:]
        gap_all = [leftover_by_year[y] for y in years]

        bar_width = 0.2
        x_all = [y for y in years]

        ax_all.bar([x - 1.5 * bar_width for x in x_all], base_demand_all, width=bar_width, label='Base Demand', color='skyblue')
        ax_all.bar([x - 0.5 * bar_width for x in x_all], injected_all, width=bar_width, label='Injected', color='lightgreen')
        ax_all.bar([x + 0.5 * bar_width for x in x_all], covered_all, width=bar_width, label='Covered by Shelf', color='orange')
        bars = ax_all.bar([x + 1.5 * bar_width for x in x_all], gap_all, width=bar_width, label='Gap', color='red')

        def add_value_labels(x_positions, values, color):
            for i, val in enumerate(values):
                if val > 0:
                    ax_all.text(x_positions[i], val + 5, str(int(val)),
                                ha='center', va='bottom', fontsize=9, fontweight='bold', color=color)

        x_base = [x - 1.5 * bar_width for x in x_all]
        x_injected = [x - 0.5 * bar_width for x in x_all]
        x_covered = [x + 0.5 * bar_width for x in x_all]
        x_gap = [x + 1.5 * bar_width for x in x_all]

        add_value_labels(x_base, base_demand_all, 'blue')
        add_value_labels(x_injected, injected_all, 'green')
        add_value_labels(x_covered, covered_all, 'orange')
        add_value_labels(x_gap, gap_all, 'red')

        ax_all.set_title("Unified View: Demand Flow and Gaps by Year", fontsize=12, weight='bold')
        ax_all.set_xlabel("Year")
        ax_all.set_ylabel("Units")
        ax_all.set_xticks(x_all)
        ax_all.set_xticklabels(all_years_str)
        ax_all.legend()
        ax_all.grid(axis='y')
        fig_all.tight_layout()

        canvas_all = FigureCanvasTkAgg(fig_all, master=result_window)
        canvas_all.draw()
        canvas_all.get_tk_widget().pack(side="bottom", pady=10, fill="both", expand=True)

        tk.Button(top_frame, text="ייצא לאקסל", command=on_export_click).pack(anchor="w", pady=5)

    except Exception as e:
        messagebox.showerror("Error", f"שגיאה בחישוב: {e}")

# ---------------- Build GUI ----------------

root = tk.Tk()
root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)
root.title("תוכנת חישוב - רכש פירוטכניקה")

# הצגת סמל תחום בצד ימין
logo_path = "סמל תחום.jpeg"
try:
    pil_logo = Image.open(logo_path).convert("RGBA")
    pil_logo.thumbnail((125, 125))
    
    # יצירת מסיכה עגולה
    mask = Image.new("L", pil_logo.size, 0)
    draw = ImageDraw.Draw(mask)
    draw.ellipse((0, 0, pil_logo.size[0], pil_logo.size[1]), fill=255)
    
    pil_logo.putalpha(mask)
    
    logo_img = ImageTk.PhotoImage(pil_logo)
    logo_label = tk.Label(root, image=logo_img)
    logo_label.image = logo_img
    logo_label.grid(row=0, column=6, padx=10, pady=10, sticky="nw")
except Exception as e:
    print(f"שגיאה בטעינת התמונה: {e}")
header_frame = tk.LabelFrame(root, text="🧾 פרטי פריט", font=("Arial", 10, "bold"))
header_frame.grid(row=0, column=0, columnspan=6, padx=5, pady=5, sticky="n")

# שורה 1: מסח"א ותיאור הפריט
row1 = tk.Frame(header_frame)
row1.pack(anchor="w", pady=2)
tk.Label(row1, text="מסח\"א:", width=12, anchor="e").pack(side="left")
entry_masach = tk.Entry(row1, width=15)
entry_masach.pack(side="left", padx=5)
tk.Label(row1, text="תיאור הפריט:", width=12, anchor="e").pack(side="left")
entry_item_desc = tk.Entry(row1, width=40)
entry_item_desc.pack(side="left", padx=5)

# שורה 2: NSN, Part Number, CAGE
row2 = tk.Frame(header_frame)
row2.pack(anchor="w", pady=2)
tk.Label(row2, text="NSN:", width=12, anchor="e").pack(side="left")
entry_nsn = tk.Entry(row2, width=15)
entry_nsn.pack(side="left", padx=5)
tk.Label(row2, text="Part Number:", width=12, anchor="e").pack(side="left")
entry_part_number = tk.Entry(row2, width=15)
entry_part_number.pack(side="left", padx=5)
tk.Label(row2, text="CAGE:", width=6, anchor="e").pack(side="left")
entry_cage = tk.Entry(row2, width=10)
entry_cage.pack(side="left", padx=5)

# שורה 3: פריטים חלופיים
row3 = tk.Frame(header_frame)
row3.pack(anchor="w", pady=2)
tk.Label(row3, text="פריטים חלופיים/סטים:", width=18, anchor="e").pack(side="left")
alt_entries = []
for _ in range(5):
    e_alt = tk.Entry(row3, width=12)
    e_alt.pack(side="left", padx=5)
    alt_entries.append(e_alt)

# שורה 4: מקדם עיתוד וניסוי אורך חיים
row4 = tk.Frame(header_frame)
row4.pack(anchor="w", pady=2)
tk.Label(row4, text="מקדם עיתוד:", width=12, anchor="e").pack(side="left")
entry_atud_factor = tk.Entry(row4, width=10)
entry_atud_factor.pack(side="left", padx=5)
tk.Label(row4, text="כמות לניסוי אורך חיים:", width=18, anchor="e").pack(side="left")
entry_life_test_qty = tk.Entry(row4, width=10)
entry_life_test_qty.pack(side="left", padx=5)
entry_life_test_qty.insert(0, "10")

row_start = 10
general_params_frame = tk.LabelFrame(root, text="⚙️ פרמטרים כלליים", font=("Arial", 10, "bold"))
general_params_frame.grid(row=row_start, column=0, columnspan=3, padx=10, pady=10, sticky="n")

tk.Label(general_params_frame, text="שנת התחלה:").grid(row=row_start, column=0)
entry_start_year = tk.Entry(general_params_frame)
entry_start_year.grid(row=row_start, column=1)
entry_start_year.insert(0, "0")

def update_year_labels(*args):
    try:
        start_yr = safe_int(entry_start_year.get())
        for i in range(9):
            year_val = start_yr + i
            year_labels[i].config(text=str(year_val))
    except:
        pass

entry_start_year.bind("<KeyRelease>", update_year_labels)

tk.Label(general_params_frame, text="אורך חיים באחסנה (חודשים):").grid(row=row_start+1, column=0)
entry_storage = tk.Entry(general_params_frame)
entry_storage.grid(row=row_start+1, column=1)
entry_storage.insert(0, "0")

tk.Label(general_params_frame, text="אורך חיים בשימוש (חודשים):").grid(row=row_start+2, column=0)
entry_use = tk.Entry(general_params_frame)
entry_use.grid(row=row_start+2, column=1)
entry_use.insert(0, "0")

tk.Label(general_params_frame, text="שנת מדף נוספת:").grid(row=row_start+3, column=0)
entry_extra_shelf_year = tk.Entry(general_params_frame)
entry_extra_shelf_year.grid(row=row_start+3, column=1)
entry_extra_shelf_year.insert(0, "0")

tk.Label(general_params_frame, text="כמות מדף נוספת:").grid(row=row_start+4, column=0)
entry_extra_shelf_qty = tk.Entry(general_params_frame)
entry_extra_shelf_qty.grid(row=row_start+4, column=1)
entry_extra_shelf_qty.insert(0, "0")

table_start = row_start + 6
header_labels = tk.Frame(root)
header_labels.grid(row=table_start, column=0, columnspan=3, pady=(0, 5), sticky="n")

tk.Label(header_labels, text="שנה", width=10).grid(row=0, column=0, padx=(5, 5))
tk.Label(header_labels, text="מורכב (פג תוקף)", width=10).grid(row=0, column=1, padx=(5, 5))
tk.Label(header_labels, text="מדף (פג תוקף)", width=10).grid(row=0, column=2, padx=(5, 5))

years_frame = tk.Frame(root)
years_frame.grid(row=table_start + 1, column=0, columnspan=3, pady=5)

installed_entries = []
shelf_entries = []
year_labels = []

for i in range(9):
    lbl = tk.Label(years_frame, text=f"(שנה {i+1})")
    lbl.grid(row=i, column=0, padx=(5, 5), sticky="e")
    year_labels.append(lbl)

    e_inst = tk.Entry(years_frame, width=10)
    e_inst.grid(row=i, column=1, padx=(5, 5))
    e_inst.insert(0, "0")
    e_inst.bind("<FocusIn>", select_all)
    installed_entries.append(e_inst)

    e_shelf = tk.Entry(years_frame, width=10)
    e_shelf.grid(row=i, column=2, padx=(5, 5))
    e_shelf.insert(0, "0")
    e_shelf.bind("<FocusIn>", select_all)
    shelf_entries.append(e_shelf)

deal_table_start = table_start + 1 + 9 + 2
deal_frame = tk.LabelFrame(root, text="📦 טבלת עסקאות", font=("Arial", 10, "bold"))
deal_frame.grid(row=deal_table_start, column=0, columnspan=6, padx=10, pady=10, sticky="n")

tk.Label(deal_frame, text="תיאור העסקה").grid(row=0, column=0)
tk.Label(deal_frame, text="כמות").grid(row=0, column=1)
tk.Label(deal_frame, text="זמן הגעה משוער").grid(row=0, column=2)

deals_desc_entries = []
deals_qty_entries = []
deals_arrival_entries = []
for i in range(3):
    e_desc = tk.Entry(deal_frame, width=15)
    e_desc.grid(row=i+1, column=0)
    deals_desc_entries.append(e_desc)

    e_qty = tk.Entry(deal_frame, width=10)
    e_qty.grid(row=i+1, column=1)
    deals_qty_entries.append(e_qty)

    e_arrival = tk.Entry(deal_frame, width=15)
    e_arrival.grid(row=i+1, column=2)
    deals_arrival_entries.append(e_arrival)

calc_button_row = deal_table_start + 2 + 4
tk.Button(root, text="חשב פער ורכש", command=calculate_gap).grid(
    row=calc_button_row, column=0, columnspan=3, pady=10
)

def on_closing():
    root.quit()
    root.destroy()

root.protocol("WM_DELETE_WINDOW", on_closing)
tk.Label(root, text="© כל הזכויות שמורות - Matan Didi", font=("Arial", 9), fg="gray").grid(
    row=calc_button_row + 2, column=0, columnspan=6, pady=10, sticky="n"
)
root.mainloop()